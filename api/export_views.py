# api/export_views.py

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from datetime import datetime
from django.utils import timezone
from decimal import Decimal
from .models import Expense, Envelope, UserValue, Goal, Debt, Tontine, TontineParticipant, Income
from .payment_views import require_premium
from django.db.models import Sum, Q
import calendar as cal_module


MOIS_FR = ['', 'Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
           'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Decembre']

CATEGORY_TO_VALUE = {
    'famille':      ['famille', 'alimentation', 'logement'],
    'spiritualite': ['spiritualite'],
    'education':    ['education'],
    'sante':        ['sante'],
    'travail':      ['transport'],
    'loisirs':      ['loisirs', 'vetements'],
    'communaute':   ['famille'],
}


def get_month_range_from_request(request):
    """
    Retourne (start_date, end_date, period_label) selon ?month=YYYY-MM
    Défaut = mois courant
    """
    month_param = request.GET.get('month')
    now = timezone.now()

    if month_param:
        try:
            year, month = map(int, month_param.split('-'))
            start = now.replace(
                year=year, month=month, day=1,
                hour=0, minute=0, second=0, microsecond=0
            )
        except (ValueError, AttributeError):
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    last_day = cal_module.monthrange(start.year, start.month)[1]
    end = start.replace(day=last_day, hour=23, minute=59, second=59)
    period_label = f"{MOIS_FR[start.month]} {start.year}"

    return start, end, period_label, now


def get_user_score(user):
    """Récupérer le score Yoonu Dal"""
    try:
        from django.test.client import RequestFactory
        from rest_framework.test import force_authenticate
        from . import views

        factory = RequestFactory()
        request = factory.get('/api/yoonu-score/')
        force_authenticate(request, user=user)
        response = views.get_yoonu_score(request)

        if hasattr(response, 'data') and response.status_code == 200:
            return int(response.data.get('total_score', 50))
    except Exception as e:
        print(f"Erreur score: {e}")
    return 50


def format_amount(amount):
    if isinstance(amount, Decimal):
        return int(round(amount))
    return int(round(float(amount)))


# ==========================================
# EXPORT EXCEL
# ==========================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_premium
def export_excel(request):
    """Export Excel — support ?month=YYYY-MM"""
    try:
        user = request.user
        start_date, end_date, period_label, now = get_month_range_from_request(request)

        print(f"\n{'='*60}")
        print(f"📊 EXPORT EXCEL - {user.username} - {period_label}")
        print(f"{'='*60}\n")

        yoonu_score = get_user_score(user)

        expenses = Expense.objects.filter(
            user=user, date__gte=start_date.date(), date__lte=end_date.date()
        ).order_by('-date')

        incomes = Income.objects.filter(
            user=user, date__gte=start_date.date(), date__lte=end_date.date()
        ).order_by('-date')

        envelopes = Envelope.objects.filter(user=user)
        goals = Goal.objects.filter(user=user, is_achieved=False)
        debts = Debt.objects.filter(user=user, is_active=True)
        tontines = TontineParticipant.objects.filter(
            user=user, tontine__status='active'
        ).select_related('tontine')
        user_values = UserValue.objects.filter(user=user).order_by('priority')[:3]

        total_expenses_raw = sum(exp.amount for exp in expenses)
        total_expenses = format_amount(total_expenses_raw)
        total_income_raw = sum(inc.amount for inc in incomes)
        total_income = format_amount(total_income_raw)
        balance = total_income - total_expenses

        wb = Workbook()

        # === SHEET 1 : Vue d'ensemble ===
        ws1 = wb.active
        ws1.title = "Vue d'ensemble"
        ws1['A1'] = f"YOONU DAL - RAPPORT {period_label.upper()}"
        ws1['A1'].font = Font(size=16, bold=True, color="10B981")
        ws1['A2'] = f"Genere le {now.strftime('%d/%m/%Y a %H:%M')}"
        ws1['A2'].font = Font(size=10, italic=True)
        ws1['A4'] = "Score Yoonu Dal"
        ws1['B4'] = yoonu_score
        ws1['A4'].font = Font(bold=True)
        ws1['B4'].font = Font(size=14, bold=True, color="10B981")
        ws1['A6'] = f"Revenus {period_label}"
        ws1['B6'] = total_income
        ws1['A7'] = f"Depenses {period_label}"
        ws1['B7'] = total_expenses
        ws1['A8'] = "Solde"
        ws1['B8'] = balance
        ws1['A9'] = "Taux d'epargne"
        ws1['B9'] = f"{(balance / total_income * 100) if total_income > 0 else 0:.1f}%"
        ws1['A10'] = "Nombre de depenses"
        ws1['B10'] = expenses.count()
        for row in range(6, 11):
            ws1[f'A{row}'].font = Font(bold=True)

        # === SHEET 2 : Revenus ===
        ws2 = wb.create_sheet("Revenus")
        ws2['A1'] = f"REVENUS DE {period_label.upper()}"
        ws2['A1'].font = Font(size=14, bold=True)
        for idx, header in enumerate(['Date', 'Source', 'Description', 'Montant'], 1):
            cell = ws2.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        for idx, income in enumerate(incomes, 4):
            ws2.cell(row=idx, column=1, value=income.date.strftime('%d/%m/%Y'))
            ws2.cell(row=idx, column=2, value=income.source or '-')
            ws2.cell(row=idx, column=3, value=income.description or '-')
            ws2.cell(row=idx, column=4, value=format_amount(income.amount)).number_format = '#,##0'
        last_row = incomes.count() + 4
        ws2[f'C{last_row}'] = "TOTAL"
        ws2[f'C{last_row}'].font = Font(bold=True)
        ws2[f'D{last_row}'] = total_income
        ws2[f'D{last_row}'].font = Font(bold=True)

        # === SHEET 3 : Dépenses ===
        ws3 = wb.create_sheet("Depenses")
        ws3['A1'] = f"DEPENSES DE {period_label.upper()}"
        ws3['A1'].font = Font(size=14, bold=True)
        for idx, header in enumerate(['Date', 'Categorie', 'Description', 'Montant'], 1):
            cell = ws3.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        for idx, expense in enumerate(expenses, 4):
            ws3.cell(row=idx, column=1, value=expense.date.strftime('%d/%m/%Y'))
            ws3.cell(row=idx, column=2, value=expense.get_category_display())
            ws3.cell(row=idx, column=3, value=expense.description or '-')
            ws3.cell(row=idx, column=4, value=format_amount(expense.amount)).number_format = '#,##0'
        last_row = expenses.count() + 4
        ws3[f'C{last_row}'] = "TOTAL"
        ws3[f'C{last_row}'].font = Font(bold=True)
        ws3[f'D{last_row}'] = total_expenses
        ws3[f'D{last_row}'].font = Font(bold=True)

        # === SHEET 4 : Enveloppes ===
        ws4 = wb.create_sheet("Enveloppes")
        ws4['A1'] = "BUDGET PAR ENVELOPPE"
        ws4['A1'].font = Font(size=14, bold=True)
        for idx, header in enumerate(['Enveloppe', 'Budget', 'Depense', 'Reste', '%'], 1):
            cell = ws4.cell(row=3, column=idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

        envelope_names = {
            'essentiels': 'Essentiels',
            'plaisirs': 'Plaisirs',
            'projets': 'Projets',
            'liberation': 'Liberation'
        }
        env_categories = {
            'essentiels': ['logement', 'alimentation', 'transport', 'sante'],
            'plaisirs': ['loisirs', 'vetements', 'autre'],
            'projets': ['education', 'famille', 'spiritualite'],
            'liberation': ['dettes']
        }

        for idx, envelope in enumerate(envelopes, 4):
            # Dépenses recalculées sur le mois demandé
            cats = env_categories.get(envelope.envelope_type, [])
            spent_real = float(
                Expense.objects.filter(
                    user=user,
                    category__in=cats,
                    date__gte=start_date.date(),
                    date__lte=end_date.date()
                ).aggregate(total=Sum('amount'))['total'] or 0
            )
            budget = float(envelope.allocated_percentage / 100) * total_income
            remaining = budget - spent_real
            percentage = int(envelope.allocated_percentage)

            ws4.cell(row=idx, column=1, value=envelope_names.get(envelope.envelope_type, envelope.envelope_type))
            ws4.cell(row=idx, column=2, value=int(budget)).number_format = '#,##0'
            ws4.cell(row=idx, column=3, value=int(spent_real)).number_format = '#,##0'
            ws4.cell(row=idx, column=4, value=int(remaining)).number_format = '#,##0'
            ws4.cell(row=idx, column=5, value=f"{percentage}%")

        # === SHEET 5 : Valeurs ===
        ws5 = wb.create_sheet("Alignement Valeurs")
        ws5['A1'] = "ALIGNEMENT AVEC TES VALEURS"
        ws5['A1'].font = Font(size=14, bold=True)
        if user_values.exists():
            for idx, header in enumerate(['Priorite', 'Valeur', 'Depenses', '% Budget', 'Cible', 'Ecart'], 1):
                cell = ws5.cell(row=3, column=idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            for idx, value in enumerate(user_values, 4):
                related_categories = CATEGORY_TO_VALUE.get(value.value, [])
                value_expenses = expenses.filter(category__in=related_categories)
                value_amount = sum(exp.amount for exp in value_expenses)
                value_percent = (float(value_amount) / float(total_expenses_raw) * 100) if total_expenses_raw > 0 else 0
                target_percent = 30 - (value.priority - 1) * 5
                gap = value_percent - target_percent
                ws5.cell(row=idx, column=1, value=value.priority)
                ws5.cell(row=idx, column=2, value=value.value)
                ws5.cell(row=idx, column=3, value=format_amount(value_amount)).number_format = '#,##0'
                ws5.cell(row=idx, column=4, value=f"{value_percent:.1f}%")
                ws5.cell(row=idx, column=5, value=f"{target_percent}%")
                ws5.cell(row=idx, column=6, value=f"{gap:+.1f}%")
        else:
            ws5['A5'] = "Complete ton diagnostic pour voir l'alignement avec tes valeurs"

        # === SHEET 6 : Objectifs ===
        ws6 = wb.create_sheet("Objectifs")
        ws6['A1'] = "OBJECTIFS EN COURS"
        ws6['A1'].font = Font(size=14, bold=True)
        if goals.exists():
            for idx, header in enumerate(['Objectif', 'Cible', 'Atteint', 'Reste', '% Progression', 'Echeance'], 1):
                cell = ws6.cell(row=3, column=idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            for idx, goal in enumerate(goals, 4):
                target = format_amount(goal.target_amount)
                current = format_amount(goal.current_amount)
                progress = (float(goal.current_amount) / float(goal.target_amount) * 100) if goal.target_amount > 0 else 0
                ws6.cell(row=idx, column=1, value=goal.title)
                ws6.cell(row=idx, column=2, value=target).number_format = '#,##0'
                ws6.cell(row=idx, column=3, value=current).number_format = '#,##0'
                ws6.cell(row=idx, column=4, value=target - current).number_format = '#,##0'
                ws6.cell(row=idx, column=5, value=f"{progress:.0f}%")
                ws6.cell(row=idx, column=6, value=goal.deadline.strftime('%d/%m/%Y') if goal.deadline else '-')
        else:
            ws6['A5'] = "Aucun objectif defini"

        # === SHEET 7 : Dettes ===
        ws7 = wb.create_sheet("Dettes")
        ws7['A1'] = "GESTION DES DETTES"
        ws7['A1'].font = Font(size=14, bold=True)
        if debts.exists():
            for idx, header in enumerate(['Dette', 'Creancier', 'Total', 'Paye', 'Reste', '% Progression'], 1):
                cell = ws7.cell(row=3, column=idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            for idx, debt in enumerate(debts, 4):
                total = format_amount(debt.total_amount)
                paid = format_amount(debt.amount_paid)
                progress = debt.progress_percentage if hasattr(debt, 'progress_percentage') else 0
                ws7.cell(row=idx, column=1, value=debt.name)
                ws7.cell(row=idx, column=2, value=debt.creditor or '-')
                ws7.cell(row=idx, column=3, value=total).number_format = '#,##0'
                ws7.cell(row=idx, column=4, value=paid).number_format = '#,##0'
                ws7.cell(row=idx, column=5, value=total - paid).number_format = '#,##0'
                ws7.cell(row=idx, column=6, value=f"{progress:.0f}%")
        else:
            ws7['A5'] = "Aucune dette active. Bravo !"

        # === SHEET 8 : Tontines ===
        ws8 = wb.create_sheet("Tontines")
        ws8['A1'] = "TONTINES ACTIVES"
        ws8['A1'].font = Font(size=14, bold=True)
        if tontines.exists():
            for idx, header in enumerate(['Tontine', 'Participants', 'Contribution', 'Position', 'Statut'], 1):
                cell = ws8.cell(row=3, column=idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            for idx, tp in enumerate(tontines, 4):
                tontine = tp.tontine
                participants_count = tontine.participants.count()
                ws8.cell(row=idx, column=1, value=tontine.name)
                ws8.cell(row=idx, column=2, value=participants_count)
                ws8.cell(row=idx, column=3, value=format_amount(tontine.monthly_contribution)).number_format = '#,##0'
                ws8.cell(row=idx, column=4, value=f"{tp.position}/{participants_count}")
                ws8.cell(row=idx, column=5, value="Actif" if tp.is_active else "Inactif")
        else:
            ws8['A5'] = "Aucune tontine active"

        # Générer fichier
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'yoonu_dal_{period_label.replace(" ", "_")}_{now.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        print(f"Export Excel OK - {period_label}\n")
        return response

    except Exception as e:
        print(f"ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Erreur: {str(e)}'}, status=500)


# ==========================================
# EXPORT PDF
# ==========================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_premium
def export_pdf(request):
    """Export PDF — support ?month=YYYY-MM"""
    try:
        user = request.user
        start_date, end_date, period_label, now = get_month_range_from_request(request)

        print(f"\n{'='*60}")
        print(f"PDF EXPORT - {user.username} - {period_label}")
        print(f"{'='*60}\n")

        yoonu_score = get_user_score(user)

        expenses = Expense.objects.filter(
            user=user, date__gte=start_date.date(), date__lte=end_date.date()
        ).order_by('-date')

        incomes = Income.objects.filter(
            user=user, date__gte=start_date.date(), date__lte=end_date.date()
        ).order_by('-date')

        envelopes = Envelope.objects.filter(user=user)
        goals = Goal.objects.filter(user=user, is_achieved=False)[:5]
        debts = Debt.objects.filter(user=user, is_active=True)
        tontines = TontineParticipant.objects.filter(
            user=user, tontine__status='active'
        ).select_related('tontine')
        user_values = UserValue.objects.filter(user=user).order_by('priority')[:3]

        total_expenses_raw = sum(exp.amount for exp in expenses)
        total_expenses = format_amount(total_expenses_raw)
        total_income_raw = sum(inc.amount for inc in incomes)
        total_income = format_amount(total_income_raw)
        balance = total_income - total_expenses
        savings_rate = (balance / total_income * 100) if total_income > 0 else 0

        response = HttpResponse(content_type='application/pdf')
        filename = f'yoonu_dal_{period_label.replace(" ", "_")}_{now.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        doc = SimpleDocTemplate(response, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=24, textColor=colors.HexColor('#10B981'),
            spaceAfter=30, alignment=1
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=16, textColor=colors.HexColor('#10B981'),
            spaceAfter=12, spaceBefore=12
        )

        # PAGE 1 : Tableau de bord
        elements.append(Paragraph("YOONU DAL", title_style))
        elements.append(Paragraph("Ton Bilan Financier Conscient", styles['Normal']))
        elements.append(Paragraph(f"{user.first_name} {user.last_name}", styles['Normal']))
        elements.append(Paragraph(f"Periode : {period_label}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph("SCORE YOONU DAL", heading_style))
        score_table = Table([[f"Score Global : {yoonu_score}/100"]], colWidths=[15*cm])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph("VUE D'ENSEMBLE", heading_style))
        overview_data = [
            ['Indicateur', 'Montant'],
            ['Revenus totaux', f'{total_income:,} FCFA'.replace(',', ' ')],
            ['Depenses totales', f'{total_expenses:,} FCFA'.replace(',', ' ')],
            ['Epargne', f'{balance:,} FCFA'.replace(',', ' ')],
            ["Taux d'epargne", f'{savings_rate:.1f}%'],
            ['Nombre de depenses', str(expenses.count())]
        ]
        overview_table = Table(overview_data, colWidths=[8*cm, 7*cm])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(overview_table)
        elements.append(PageBreak())

        # PAGE 2 : Valeurs
        elements.append(Paragraph("ALIGNEMENT AVEC TES VALEURS", heading_style))
        if user_values.exists():
            values_data = [['Valeur', 'Depenses', '% Budget', 'Score Alignement']]
            for value in user_values:
                related_categories = CATEGORY_TO_VALUE.get(value.value, [])
                value_expenses_qs = expenses.filter(category__in=related_categories)
                value_amount = sum(exp.amount for exp in value_expenses_qs)
                value_percent = (float(value_amount) / float(total_expenses_raw) * 100) if total_expenses_raw > 0 else 0
                target_percent = 30 - (value.priority - 1) * 5
                alignment = max(0, min(100, 100 - abs(value_percent - target_percent) * 3))
                values_data.append([
                    value.value,
                    f'{format_amount(value_amount):,} F'.replace(',', ' '),
                    f'{value_percent:.1f}%',
                    f'{alignment:.0f}%'
                ])
            values_table = Table(values_data, colWidths=[5*cm, 4*cm, 3*cm, 3*cm])
            values_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(values_table)
        else:
            elements.append(Paragraph("Complete ton diagnostic pour voir l'alignement.", styles['Normal']))
        elements.append(PageBreak())

        # PAGE 3 : Enveloppes
        elements.append(Paragraph("SYSTEME DES 4 ENVELOPPES", heading_style))
        env_categories = {
            'essentiels': ['logement', 'alimentation', 'transport', 'sante'],
            'plaisirs': ['loisirs', 'vetements', 'autre'],
            'projets': ['education', 'famille', 'spiritualite'],
            'liberation': ['dettes']
        }
        envelope_names = {
            'essentiels': 'Essentiels',
            'plaisirs': 'Plaisirs',
            'projets': 'Projets',
            'liberation': 'Liberation'
        }
        if envelopes.exists():
            env_data = [['Enveloppe', 'Budget', 'Depense', 'Reste', '% Utilise']]
            for env in envelopes:
                cats = env_categories.get(env.envelope_type, [])
                spent_real = float(
                    Expense.objects.filter(
                        user=user, category__in=cats,
                        date__gte=start_date.date(), date__lte=end_date.date()
                    ).aggregate(total=Sum('amount'))['total'] or 0
                )
                budget = float(env.allocated_percentage / 100) * total_income
                remaining = budget - spent_real
                percent_used = (spent_real / budget * 100) if budget > 0 else 0
                env_data.append([
                    envelope_names.get(env.envelope_type, env.envelope_type),
                    f'{int(budget):,} F'.replace(',', ' '),
                    f'{int(spent_real):,} F'.replace(',', ' '),
                    f'{int(remaining):,} F'.replace(',', ' '),
                    f'{percent_used:.0f}%'
                ])
            env_table = Table(env_data, colWidths=[4*cm, 3.5*cm, 3.5*cm, 3*cm, 2*cm])
            env_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(env_table)
        elements.append(PageBreak())

        # PAGE 4 : Objectifs
        elements.append(Paragraph("OBJECTIFS EN COURS", heading_style))
        if goals.exists():
            goals_data = [['Objectif', 'Cible', 'Atteint', 'Reste', '% Progression']]
            for goal in goals:
                target = format_amount(goal.target_amount)
                current = format_amount(goal.current_amount)
                progress = (float(goal.current_amount) / float(goal.target_amount) * 100) if goal.target_amount > 0 else 0
                goals_data.append([
                    goal.title[:25],
                    f'{target:,} F'.replace(',', ' '),
                    f'{current:,} F'.replace(',', ' '),
                    f'{target - current:,} F'.replace(',', ' '),
                    f'{progress:.0f}%'
                ])
            goals_table = Table(goals_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3*cm])
            goals_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(goals_table)
        else:
            elements.append(Paragraph("Aucun objectif defini pour le moment.", styles['Normal']))
        elements.append(PageBreak())

        # PAGE 5 : Dettes
        elements.append(Paragraph("GESTION DES DETTES", heading_style))
        if debts.exists():
            total_debt = sum(d.total_amount for d in debts)
            total_paid_debt = sum(d.amount_paid for d in debts)
            debt_overview = [
                ['Indicateur', 'Montant'],
                ['Total dettes', f'{format_amount(total_debt):,} F'.replace(',', ' ')],
                ['Deja paye', f'{format_amount(total_paid_debt):,} F'.replace(',', ' ')],
                ['Reste a payer', f'{format_amount(total_debt - total_paid_debt):,} F'.replace(',', ' ')],
            ]
            debt_overview_table = Table(debt_overview, colWidths=[8*cm, 7*cm])
            debt_overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F97316')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(debt_overview_table)
            elements.append(Spacer(1, 0.5*cm))
            debts_detail = [['Dette', 'Total', 'Paye', '% Progression']]
            for debt in debts:
                progress = debt.progress_percentage if hasattr(debt, 'progress_percentage') else 0
                debts_detail.append([
                    debt.name[:25],
                    f'{format_amount(debt.total_amount):,} F'.replace(',', ' '),
                    f'{format_amount(debt.amount_paid):,} F'.replace(',', ' '),
                    f'{progress:.0f}%'
                ])
            debts_table = Table(debts_detail, colWidths=[6*cm, 3*cm, 3*cm, 3*cm])
            debts_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F97316')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(debts_table)
        else:
            elements.append(Paragraph("Aucune dette active. Bravo !", styles['Normal']))
        elements.append(PageBreak())

        # PAGE 6 : Tontines
        elements.append(Paragraph("TONTINES ACTIVES", heading_style))
        if tontines.exists():
            tontines_data = [['Tontine', 'Participants', 'Contribution', 'Cycle', 'Statut']]
            for tp in tontines:
                tontine = tp.tontine
                participants_count = tontine.participants.count()
                tontines_data.append([
                    tontine.name[:25],
                    str(participants_count),
                    f'{format_amount(tontine.monthly_contribution):,} F'.replace(',', ' '),
                    f"{tp.position}/{participants_count}",
                    "Actif" if tp.is_active else "Inactif"
                ])
            tontines_table = Table(tontines_data, colWidths=[5*cm, 2.5*cm, 3*cm, 2.5*cm, 3*cm])
            tontines_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(tontines_table)
        else:
            elements.append(Paragraph("Aucune tontine active.", styles['Normal']))
        elements.append(PageBreak())

        # PAGE 7 : Top dépenses
        elements.append(Paragraph("TOP 10 DEPENSES", heading_style))
        top_expenses = expenses.order_by('-amount')[:10]
        if top_expenses.exists():
            expenses_data = [['Description', 'Categorie', 'Montant', 'Date']]
            for exp in top_expenses:
                amount = format_amount(exp.amount)
                expenses_data.append([
                    exp.description[:25] if exp.description else '-',
                    exp.get_category_display()[:15],
                    f'{amount:,} F'.replace(',', ' '),
                    exp.date.strftime('%d/%m/%Y')
                ])
            expenses_table = Table(expenses_data, colWidths=[5*cm, 4*cm, 3*cm, 3*cm])
            expenses_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(expenses_table)

        doc.build(elements)
        print(f"Export PDF OK - {period_label}\n")
        return response

    except Exception as e:
        print(f"ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return Response({'error': f'Erreur: {str(e)}'}, status=500)
